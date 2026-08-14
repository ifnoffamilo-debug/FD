from __future__ import annotations

from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

TX_HEADERS=["ID","Тип","Сумма","Категория","Объект","Авто","Тип авторасхода","Топливо","Литры","Пробег","Комментарий","Чек","Дата","Источник"]


def _sheet_for_rows(ws, rows):
    ws.append(TX_HEADERS)
    for c in ws[1]: c.font=Font(bold=True)
    for r in rows:
        ws.append([
            r["id"],"Доход" if r["tx_type"]=="income" else "Расход",float(r["amount"]),r["category"],
            r.get("object_name") or "",r.get("vehicle_name") or "",r.get("vehicle_expense_type") or "",
            r.get("fuel_type") or "",r.get("fuel_liters") or "",r.get("odometer") or "",r.get("comment") or "",
            "Да" if r.get("receipt_file_id") else "",r["created_at"],r["source"]
        ])
    ws.freeze_panes="A2";ws.auto_filter.ref=ws.dimensions
    widths=[8,12,15,24,28,24,22,14,12,14,40,10,20,14]
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
    for c in ws["C"][1:]: c.number_format='#,##0.00 "₽"'


async def create_excel(rows:list[dict], objects:list[dict], vehicles:list[dict], debts:list[dict], debt_payments:list[dict], export_dir:Path)->Path:
    export_dir.mkdir(parents=True,exist_ok=True);path=export_dir/f"fd_finance_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    wb=Workbook();wb.remove(wb.active)
    for scope,title in (("work","Рабочие финансы"),("personal","Личные финансы")):
        ws=wb.create_sheet(title);_sheet_for_rows(ws,[r for r in rows if r.get("finance_scope","work")==scope])
    ws=wb.create_sheet("Объекты");ws.append(["ID","Объект","Цена заказа","Бюджет расходов","Статус"])
    for r in objects: ws.append([r["id"],r["name"],r["contract_amount"],r["budget_amount"],r["status"]])
    ws=wb.create_sheet("Автомобили");ws.append(["ID","Название","Пробег","Статус"])
    for r in vehicles: ws.append([r["id"],r["name"],r["current_mileage"],r["status"]])
    ws=wb.create_sheet("Долги");ws.append(["ID","Тип","Кто","Исходная сумма","Погашено","Остаток","Срок","Статус","Комментарий","Создан"])
    for r in debts:
        ws.append([r["id"],"Мне должны" if r["direction"]=="to_me" else "Я должен",r["person"],float(r["original_amount"]),float(r.get("paid_amount") or 0),float(r.get("remaining") or 0),r.get("due_date") or "",r["status"],r.get("comment") or "",r["created_at"]])
    for col in ("D","E","F"):
        for c in ws[col][1:]:c.number_format='#,##0.00 "₽"'
    ws=wb.create_sheet("Погашения долгов");ws.append(["ID","Долг ID","Тип","Кто","Сумма","Дата","Операция в финансах"])
    for r in debt_payments:
        ws.append([r["id"],r["debt_id"],"Мне должны" if r["direction"]=="to_me" else "Я должен",r["person"],float(r["amount"]),r["paid_at"],r.get("transaction_id") or ""])
    for c in ws["E"][1:]:c.number_format='#,##0.00 "₽"'
    wb.save(path);return path
